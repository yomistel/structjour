'''
Created on Oct 18, 2018

@author: Mike Petersen
'''
import os
import pandas as pd
import numpy as np
import datetime as dt

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, colors

# from inspiration.inspire import Inspire
from inspiration.inspire import Inspire
from journal.dfutil import DataFrameUtil
from journal.definetrades import FinReqCol
from journal.xlimage import XLImage
from journal.tradestyle import c as tcell
from journal.tradestyle import style_range
from journal.thetradeobject import TheTradeObject, SumReqFields

# pylint: disable=C0103, C0201, W0703

def askUser(question):
    '''
    Ask the user a question. Placed in a function to facilitate automating it.
    :return: The response
    '''
    response = input(question)
    return response


class LayoutSheet:
    '''
    Contains methods to layout the material on the excel page. Uses both
    pandas and openpyxl depending on how far in the program.  Generally
    the program progresses from pandas to openpyxl.
    '''

    def __init__(self, topMargin, inputlen, spacing=3):
        '''
        Constructor
        :params topMargin: The space at the top before the trade table. Includes the
                           inspire quote and space notes.
        :params inputlen: Don't enter a value here. Its the length of the dframe after processing.
        :params spacing: The space between trade summaries.
        :params frq: The FinReqCol object
        '''
        srf = SumReqFields()
        sumSize = srf.maxrow() + 5
        self.summarySize = sumSize
        self.topMargin = topMargin
        self.inputlen = inputlen
        self.spacing = spacing
        self.DSFAnchor = None

    def imageData(self, df, ldf, ft="png"):
        '''
        Gather the image names and determine the locations in the Excel doc to place them. Excel
        has a few things at top followed by trade summaries, charts and tables for each trade.
        Return with the image name/location data structure. The structure can be used for the Excel
        DataFrame-- to navigate summary form locations and just for the names
        :params df: The DataFrame representing the input file plus some stuff added in
                    processOutputFile
        :params ldf: A list of dataFrames. Each encapsulates a trade.
        :parmas ft: Image filetype extension. (NOT USED)
        :return (Imagelocation, df): ImageLocation contains information about the excel document
                    locations of trade summaries and image locations. The dataFrame df is the
                    outline used to create the workbook, ImageLocation will be used to stye it
                    and fill in the stuff.
        '''
        # Add rows and append each trade, leaving space for an image. Create a list of
        # names and row numbers to place images within the excel file (imageLocation
        # data structure).

        # Number of rows between trade summaries
        frq = FinReqCol()
        newdf = DataFrameUtil.createDf(df, self.topMargin)

        df = newdf.append(df, ignore_index=True)
 
        imageLocation = list()
        count = 0
        for tdf in ldf:
            imageName = '{0}_{1}_{2}_{3}.{4}'.format(tdf[frq.tix].unique()[-1].replace(' ', ''),
                                                     tdf[frq.name].unique()[-1].replace(' ', '-'),
                                                     tdf[frq.start].unique()[-1],
                                                     tdf[frq.dur].unique()[-1], ft)

            # Holds location, deprected name, image name, trade start time, trade duration as delta
            imageLocation.append([len(tdf) + len(df) + self.spacing,
                                  tdf.Tindex.unique()[0].replace(
                                      ' ', '') + '.' + ft,
                                  imageName,
                                  tdf.Start.unique()[-1],
                                  tdf.Duration.unique()[-1]])
            # print(count, imageName, len(imageLocation), len(tdf) + len(df) + 3)
            count = count + 1

            # Append the mini trade table then add rows to fit the tradeSummary form
            df = df.append(tdf, ignore_index=True)
            df = DataFrameUtil.addRows(df, self.summarySize)
        return imageLocation, df

    def createWorkbook(self, dframe):
        '''
        Create the workbook obj and give it all the data in the DataFrame. This copies
        almost verbatim each cell in the DataFrame to a cell in the workbook--except we add the
        column headers from the DataFrame to the correct location above the table in
        row [self.topMargin].
        :params dframe: The trades and summaries already formatted in the correct shape for this
                         new document we are creating.
        :return (wb, ws, nt): The workbook, its worksheet and the original DataFrame
        '''
        nt = dframe
        # def
        wb = Workbook()
        ws = wb.active

        # Add all cell values from the df to the ws object
        for r in dataframe_to_rows(nt, index=False, header=False):
            ws.append(r)

        # Place column names at the top table -- (under the notes and inspire quote)
        for name, cell in zip(nt.columns, ws[self.topMargin]):
            cell.value = name
        return wb, ws, nt

    def styleTop(self, ws, widthDF, tf):
        '''
        Style the table, and the top bit. Here we style the table and the things above it. The
        table data is already there. Above the table are two merged groups. At the top is the
        inspire quote. Next is an introductory notes section. There is currently no external
        control to set the sizes of these two things. Its hard coded here here. 
        :params ws: The openpyxl Worksheet to use.
        :params widthDF: The width of the dataFrame holding the trades.
        :params tf: The TradeFormat object
        :prerequisites: Depends that ws has the table data beginning A25 (1,25), beginning with
                        column headers on row 25. The length of the table is defined in
                        self.inputlen. If topMargin changes, this code will need change.
        '''
        # Hard coded sizes here
        quoteRange = [(1, 1), (13, 5)]
        noteRange = [(1, 6), (13, 24)]
    
        tblRng = "{0}:{1}".format(tcell((1, self.topMargin)), tcell(
            (widthDF, self.topMargin + self.inputlen)))
        tab = Table(displayName="Table1", ref=tblRng)
        style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style

        ws.add_table(tab)

        # A1:M5 inspire quote. Style the quote and insert the quote
        tf.mergeStuff(ws, quoteRange[0], quoteRange[1])
        ws["A1"].style = tf.styles["normStyle"]
        style_range(ws, "A1:M5", border=tf.styles["normStyle"].border)
        inspire = Inspire()
        ws["A1"] = inspire.getrandom().replace("\t", "        ")

        # A6:M24 introductory notes. Style the cells and enter minimal text.
        tf.mergeStuff(ws, noteRange[0], noteRange[1])
        ws["A6"].style = tf.styles["explain"]
        style_range(ws, "A6:M24", border=tf.styles["explain"].border)

    def runSummaries(self, imageLocation, ldf, jf, ws, tf):
        '''
        This is a runner script. For each trade DataFrame in the list ldf we will get and place
        the chart image, call TheTradeObject.runSummary to gather the summary data into the
        TradeObject. Then we will create the trade summaries, by styling the form and placing the
        summary data/forms next to the images we just placed.
        :params imageLocation: Data structure containing the locations to place summaries
        :params ldf: A list of DataFrames, each representing a single trade with one or more
                     transactions
        :params jf: The JournalFiles object containing needed path locations
        :params ws: The openpyx Worksheet object to work on
        :params tf: The TradeFormat object with data and methods for creating the Trade Summaries.
        :return tradeSummaries: A list of 1 row DataFrames created by TheTradeObject. Each has 1
                    row representing one trade and contains multiple columns for entries and exits.
        '''
        tradeSummaries = list()
        XL = XLImage()
        srf = SumReqFields()

        response = askUser("Would you like to enter strategy names, targets and stops?   ")
        interview = True if response.lower().startswith('y') else False

        for loc, tdf in zip(imageLocation, ldf):

            img = XL.getAndResizeImage(loc[2], jf.outdir)

            # Hidden here is the location to place the chart on the page.
            if img:
                col = srf.maxcol() + 1
                col = tcell((col, 1))[0]
                cellname = col + str(loc[0])
                ws.add_image(img, cellname)

            #Put together the trade summary info for each trade and interview the trader
            tto = TheTradeObject(tdf, interview, srf)
            tto.runSummary(None)
            tradeSummaries.append(tto.TheTrade)

            #Place the format shapes/styles in the worksheet
            tf.formatTrade(ws, srf, anchor=(1, loc[0]))

            #populate the trade information
            for key in srf.tfcolumns.keys():
                cell = srf.tfcolumns[key][0]
                if isinstance(cell, list):
                    cell = cell[0]
                tradeval = tto.TheTrade[key].unique()[0]
                # print ("{0:10} \t{3} \t{1:}\t{2} ".format(key, cell, tradeval,
                # tcell(cell, anchor=(1, loc[0]))))

                # Put some formulas in each trade Summary
                if key in srf.tfformulas:

                    anchor = (1, loc[0])
                    formula = srf.tfformulas[key][0]
                    args = []
                    for c in srf.tfformulas[key][1:]:
                        args.append(tcell(c, anchor=anchor))
                    tradeval = formula.format(*args)

                if not tradeval:
                    continue
                if isinstance(tradeval, (pd.Timestamp, dt.datetime, np.datetime64)):
                    tradeval = pd.Timestamp(tradeval)


                ws[tcell(cell, anchor=(1, loc[0]))] = tradeval

        # print("Done with interview")
        return tradeSummaries

    def populateMistakeForm(self, tradeSummaries, mistake, ws, imageLocation):
        '''
        Populate the dynamic parts of mistake summaries. That includes fomulas with references to
        tradeSummaries and hyperlinks to the same. The anchor info for the tradeSummaries cell translation is in
        imageLocation and the specific location of the transaltions is in the mistakeFields. Th
        return hyperlinks in the tradeSummaries forms are also translated here.
        The form and the static content are already created presumably in the ws in the arguments.
        Either way we create the info into the given ws.
        :Programming note: MistakeSummaries form is not a one-place creation because of the hard-
                           coded stuff in this method. With some careful work, it could be. It
                           would involve placing all the data we need for the hyperlinks (other
                           than cell transation), in the mistakeFields or in the formula dict.
                           blwntffczmlz
        :params tradeSummaries: A dataframe containing the the trade summaries info, one line per
                                trade.
        :parmas mistake: A dataframe containing the info to populate the mistake summary form.
        :params ws: The openpyxl worksheet object.
        :parmas imageLocation: A list containing the locations in the worksheet for each of the
                               trades in tradeSummaries.
        '''

        # Populate the name fields as hyperlinks to tradeSummaries title cell and back.
        for i, (iloc, tsum) in enumerate(zip(imageLocation, tradeSummaries)):
            key = "name" + str(i+1)
            cell = mistake.mistakeFields[key][0][0]
            cell = tcell(cell, anchor=mistake.anchor)
            targetcell = (1, iloc[0])
            targetcell = tcell(targetcell)
            cellval = "{0} {1} {2}".format(
                i+1, tsum.Name.unique()[0], tsum.Account.unique()[0])
            link = "#{}!{}".format(ws.title, targetcell)

            ws[cell].hyperlink = (link)
            ws[cell] = cellval
            ws[cell].font = Font(color=colors.WHITE, underline="double")

            link = "#{}!{}".format(ws.title, cell)
            ws[targetcell].hyperlink = (link)
            ws[targetcell].font = Font(
                color=colors.WHITE, size=16, underline="double")
                        


        # Populate the pl (loss) fields and the mistake fields. These are all simple formulas
        # like =B31
        tokens = ["tpl", "pl", "mistake"]
        for token in tokens:
            for i in range(len(tradeSummaries)):
                key = token + str(i+1)
                if isinstance(mistake.mistakeFields[key][0], list):
                    cell = mistake.mistakeFields[key][0][0]
                else:
                    cell = cell = mistake.mistakeFields[key][0]
                cell = tcell(cell, anchor=mistake.anchor)
            #     print(cell)
                formula = mistake.formulas[key][0]
                targetcell = mistake.formulas[key][1]
                targetcell = tcell(targetcell, anchor=(1, imageLocation[i][0]))
                formula = formula.format(targetcell)

                # print("ws[{0}]='{1}'".format(cell, formula))
                ws[cell] = formula

    def populateDailySummaryForm(self, TheTradeList, mistake, ws, anchor):
        '''
        Populate the daily Summary Form. The PL values are retrieved from TheTradeList. The static
        labels are set earlier. This method sets some statistics and notes for things like
        regarding average winners/losers etc.
        :params listOfTrade: A python list of the Summary Trade DataFrame, aka TheTrade, each one
                             is a single row DataFrame containg all the data for trade summaries.
        :params mistke:
        :params ws: The openpyxl Worksheet object
        :raise Value Error: When pl is misformatted and cannot be used.
        '''
        srf = SumReqFields()
        liveWins = list()
        liveLosses = list()
        simWins = list()
        simLosses = list()
        maxTrade = (0, "notrade")
        minTrade = (0, "notrade")
        #Didnot save the Trade number in TheTrade.  These should be the same order...
        count = 0

        for TheTrade in TheTradeList:
            pl = TheTrade[srf.pl].unique()[0]
            live = True if TheTrade[srf.acct].unique()[0] == "Live" else False
            count = count + 1

            # A bug-ish inspired baby-sitter
            if isinstance(pl, str):
                if pl == '':
                    pl = 0
                else:
                    try:
                        pl = float(pl)
                    except NameError:
                        raise ValueError(
                            'Malformed float for variable pl in createDailySummary')

            # print(pl)
            if float(pl) > maxTrade[0]:
                maxTrade = (pl, "Trade{0}, {1}, {2}".format(
                    count, TheTrade[srf.acct].unique()[0], TheTrade[srf.name].unique()[0]))
            if pl < minTrade[0]:
                minTrade = (pl, "Trade{0}, {1}, {2}".format(
                    count, TheTrade[srf.acct].unique()[0], TheTrade[srf.name].unique()[0]))

            if live:
                if pl > 0:
                    liveWins.append(pl)
                else:
                    liveLosses.append(pl)
            else:
                if pl > 0:
                    simWins.append(pl)
                else:
                    simLosses.append(pl)

        anchor = (anchor[0], anchor[1] + mistake.numTrades + 5)
        self.DSFAnchor = anchor

        dailySumData = dict()
        dailySumData['livetot'] = sum([sum(liveWins), sum(liveLosses)])

        numt = len(liveWins) + len(liveLosses)
        if numt == 0:
            dailySumData['livetotnote'] = "0 Trades"
        else:
            note = "{0} Trade{1}, {2} Winner{3}, {4}, Loser{5}"
            note = note.format(numt, "" if numt == 1 else "s", len(liveWins),
                               "" if len(liveWins) == 1 else "s", len(
                                   liveLosses),
                               "" if len(liveLosses) == 1 else "s")
            dailySumData['livetotnote'] = note
        dailySumData['simtot'] = sum([sum(simWins), sum(simLosses)])

        # 9 trades,  3 Winners, 6 Losers
        numt = len(simWins) + len(simLosses)
        if numt == 0:
            dailySumData['simtotnote'] = "0 Trades"
        else:  # 4 trades, 1 Winner, 3 Losers
            note = "{0} Trade{1}, {2} Winner{3}, {4}, Loser{5}"
            note = note.format(numt, "" if numt == 1 else "s",
                               len(simWins), "" if len(simWins) == 1 else "s",
                               len(simLosses), "" if len(simLosses) == 1 else "s")
            dailySumData['simtotnote'] = note

        dailySumData['highest'] = maxTrade[0]
        dailySumData['highestnote'] = maxTrade[1]
        dailySumData['lowest'] = minTrade[0]
        dailySumData['lowestnote'] = minTrade[1]
        if (len(liveWins) + len(simWins)) == 0:
            dailySumData['avgwin'] = 0
        else:
            dailySumData['avgwin'] = sum(
                [sum(liveWins), sum(simWins)]) / (len(liveWins) + len(simWins))
        dailySumData['avgwinnote'] = "X {} =  ${:.2f}".format(
            len(liveWins) + len(simWins), sum([sum(liveWins), sum(simWins)]))
        if len(liveLosses) + len(simLosses) == 0:
            dailySumData['avgloss'] = 0
        else:
            dailySumData['avgloss'] = sum([sum(liveLosses), sum(
                simLosses)]) / (len(liveLosses) + len(simLosses))
        dailySumData['avglossnote'] = "X {} =  (${:.2f})".format(
            len(liveLosses) + len(simLosses), abs(sum([sum(liveLosses), sum(simLosses)])))

        for key in dailySumData.keys():
            rng = mistake.dailySummaryFields[key][0]
            if isinstance(rng, list):
                rng = rng[0]
            ws[tcell(rng, anchor=anchor)] = dailySumData[key]

    def save(self, wb, jf):
        '''
        Save wb as an excel file. If Permission Denied error is thrown, try renaming it.
        '''
        #Write the file
        jf.mkOutdir()
        saveName = jf.outpathfile
        count = 1
        while True:
            try:
                wb.save(saveName)
            except PermissionError as ex:
                print()
                print(ex)
                print()
                print("Failed to create file {0}.{1}".format(saveName, ex))
                print(
                    "Images from the clipboard were saved  in {0}".format(jf.outdir))
                (nm, ext) = os.path.splitext(jf.outpathfile)
                saveName = "{0}({1}){2}".format(nm, count, ext)
                print("Will try to save as {0}".format(saveName))
                count = count+1
                if count == 6:
                    print("Giving up. PermissionError")
                    raise (PermissionError(
                        "Failed to create file {0}".format(saveName)))
                continue
            except Exception as ex:
                print(ex)
            break